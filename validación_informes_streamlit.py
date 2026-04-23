# -*- coding: utf-8 -*-
import io
import os
import re
import zipfile
import unicodedata
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Optional, Tuple, List, Dict
import math
from PIL import Image
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================
# --------- CONFIG UI --------
# ============================
LOGO_PATH = Path(__file__).parent / "logo.png"
logo_img = Image.open(LOGO_PATH)
st.set_page_config(page_title="Validaciones Informes MER VIHCA", page_icon=logo_img, layout="wide")

c_logo, c_title = st.columns([1, 9])
with c_logo:
    st.image(logo_img, width=300)
with c_title:
    st.title("✅ Portal de validación de indicadores MER (VIHCA)")

def render_footer(org="VIHCA / M&E Regional", app_name="Portal de Validación MER", version="v1.0"):
    year = datetime.now().year
    st.markdown(
        f"""
        <style>
            .footer {{
                position: fixed;
                left: 0;
                bottom: 0;
                width: 100%;
                background: rgba(120,255,255,0.92);
                border-top: 1px solid rgba(0,0,0,0.08);
                padding: 10px 18px;
                text-align: center;
                font-size: 12px;
                color: #6b7280;
                z-index: 9999;
                backdrop-filter: blur(6px);
            }}
            .footer b {{
                color: #111827;
            }}
            /* Para que el contenido no quede tapado por el footer */
            .block-container {{
                padding-bottom: 70px !important;
            }}
        </style>
        <div class="footer">
            © {year} <b>{org}</b> — {app_name} {version}. Todos los derechos reservados.
        </div>
        """,
        unsafe_allow_html=True,
    )

# Llamada (una vez)
render_footer(org="Proyecto VIHCA", app_name="Validaciones Informes MER VIHCA", version="v1.0.0")

def _build_doc_md() -> str:
    return r"""
# 📖 Documentación – Validaciones de indicadores MER

## 1. Introducción
- Este portal describe la estructura y funcionalidad del portal de validaciones automáticas para los indicadores del proyecto VIHCA,
  conforme a los **lineamientos establecidos por la Guía MER de PERFAR.** 
- De manera que esto permita asegurar la integridad, consistencia y calidad de los datos reportados, mediante la ejecución de diagnósticos automatizados y presentación de alertas tempranas de errores encontrados para su oportuna corrección. 

## 2. Instrucciones de uso del Portal de Validaciones 
1. Ingresar al portal mediante el enlace de conexión compartido por el equipo de M&E regional.
2. **Cargar** uno o varios `.xlsx` o un `.zip` (con subcarpetas).
3. Pulsar **Procesar**.
4. Usar **Segmentadores** (País / Departamento / Sitio).
5. Revisar **Resumen**, **% de error**, **Detalle**, **Métricas**.
6. **Descargar** Excel completo o filtrado.
7. Aplicar los cambios necesarios para corregir o actualizar el dato alertado, antes del envío final a su jefatura inmediata.

## 3. Objetivos del Portal de Validaciones

- Detectar errores comunes de forma anticipada, en las bases de datos locales de cada país, antes de cargar datos en DATIM.
- Generar visualizaciones y una tabla resumen de los errores encontrados de los archivos cargados en el Portal.
- Fortalecer la calidad y confiabilidad de los datos reportados por los equipos en cada país. 



## 3. Indicadores y reglas que se validan
- **Formato fecha diagnóstico (HTS_TST)**
  - Regla: Utilizar formato de fehca `dd/mm/yyyy`.
- **ID duplicado (HTS_TST)**
  - Regla: Se verifica que los registros de las pruebas no se repita el mismo ID en el trimestre.
- **Fecha de inicio de TARV < Fecha del diagnóstico (HTS_TST)**
  - Regla: La `Fecha inicio TARV`no debe ser menor que la `Fecha del diagnóstico`.
- **CD4 vacío en diagnósticos positivos (HTS_TST)**
  - Regla: Se verifica que si el `Resultado de la prueba es = Positivo`, el campo de `CD4 Basal` no debe estar vacío.
- **TX_PVLS Numerador > TX_PVLS Denominador**
  - Regla: Se verifica que el `Numerador` no sea  mayor que el `Denominador`.
  - Variables que se revisan: **Sexo + Tipo de población + Rango de edad**.
- **TX_PVLS Denominador > TX_CURR**
  - Regla: Se verifica que el `Denominador` (TX_PVLS) no se mayor que el `TX_CURR`.
  - Variables que se revisan: **Sexo + Tipo de población + Rango de edad**.
- **TX_CURR ≠ Dispensación_TARV (cuadros dentro de TX_CURR)**
  - Regla: Se verifica que el valor por sexo y rango de edad sea el mismo valor en ambos cuadros
  - Variables que se revisan **Sexo y Rango de edad**.
- **Verificación de Sexo (HTST)**
  - Regla: Se verifica que en la columna Sexo, solo se registre `Femenino` o `Masculino`, si no vienen estas dos opciones el sistema detecta como error.
  ## 4. Segmentadores (filtros)
- En esta sección podrá seleccionar:
    - Orden: **País** → **Departamentos** → **Sitios**.

## 5. Cálculos y % de errores
- **Errores**: cantidad de errores encontrados.
- **% Error** = `errores / chequeos * 100`.


## 6. Archivo exportable Excel
- Hojas:
  - **Resumen** (Número de errores encontrados por indicador).
  - **Resumen de errores encontrados por indicador (en hojas separadas)**.`De no encontrarse errores no se mostrará la hoja en el archivo.`
  - Se resalta en rojo la **columna con error** en cada hoja.

## 7. Recomendaciones 
- Cada error identificado de manera automatizada permitirá fortalecer y mejorar la capacitación del dato en campo, en las clínicas o durante el procesamiento de las bases de datos.  
- Con base en la frecuencia de errores encontrados, podrán reforzar las indicaciones y el procedimiento sobre cómo se construye un indicador según la Guía MER.   
  Puede que no existan **checks** válidos en esa selección; revisa filtros/fechas.
- Mantener un registro histórico de los errores encontrados más frecuentes y tener documentado las acciones correctivas respaldará el seguimiento oportuno de cada país para asegurar la calidad del dato. 
"""

with st.expander("📖 Documentación (clic para ver)", expanded=False):
    _DOC_MD = _build_doc_md()
    st.markdown(_DOC_MD)
    st.download_button(
        "⬇️ Descargar documentación",
        _DOC_MD.encode("utf-8"),
        file_name="documentacion_validaciones.md",
        mime="text/markdown",
        use_container_width=True,
    )

col_u1, col_u2 = st.columns([3, 2])
with col_u1:
    subida_multiple = st.file_uploader(
        "📂 Cargar .xlsx (varios) o 1 .zip con subcarpetas",
        type=["xlsx", "zip"],
        accept_multiple_files=True
    )
with col_u2:
    default_pais = st.text_input("País por defecto", "Desconocido")
    default_mes = st.text_input("Mes por defecto ", "Desconocido")

procesar = st.button("▶️ Procesar", use_container_width=True)

# ============================
# ---- ESTADO (CACHE/STORE) --
# ============================
def _ss_default(key, val):
    if key not in st.session_state: st.session_state[key] = val

for key, val in {
    "processed": False,
    "df_num": pd.DataFrame(),      # Numerador > Denominador
    "df_txpv": pd.DataFrame(),     # Denominador > TX_CURR
    "df_cd4": pd.DataFrame(),      # CD4 vacío positivo
    "df_tarv_gt": pd.DataFrame(),  # TARV inicio < Diagnóstico
    "df_fdiag": pd.DataFrame(),    # Formato fecha diagnóstico
    "df_currq": pd.DataFrame(),    # TX_CURR ≠ Dispensación_TARV
    "df_iddup": pd.DataFrame(),    # ID (expediente) duplicado
    "df_sexo": pd.DataFrame(),     # Sexo inválido (HTS_TST)
    "df_txml_cita": pd.DataFrame(),  # <-- TX_ML
    "df_txcurr_cohorte": pd.DataFrame(),  # <-- Conciliación TX_CURR
    "df_txcurr_auditoria": pd.DataFrame(),  # <-- Auditoría tipo Auditoria_Sitio
    "metrics_global": defaultdict(lambda: {"errors": 0, "checks": 0}),
    "metrics_by_pds": defaultdict(lambda: {"errors": 0, "checks": 0}),
}.items():
    _ss_default(key, val)

_ss_default("sel_pais", "Todos")
_ss_default("sel_depto", "Todos")
_ss_default("sel_sitio", "Todos")

# ============================
# ----- CONSTANTES / HELPERS -
# ============================
IND_NUM_GT_DEN      = "num_gt_den"
IND_DEN_GT_CURR     = "den_gt_curr"
IND_CD4_MISSING     = "cd4_missing"
IND_TARV_GT_DIAG    = "tarv_gt_diag"     # Fecha inicio TARV < Fecha diagnóstico
IND_DIAG_BAD_FMT    = "diag_bad_format"
IND_CURR_Q1Q2_DIFF  = "curr_q1q2_diff"    # TX_CURR ≠ Dispensación_TARV
IND_ID_DUPLICADO    = "id_duplicado"      # ID (expediente) duplicado
IND_SEXO_INVALID    = "sexo_invalid"      # Sexo inválido (HTS_TST)

# --- Indicador TX_ML
IND_TXML_CITA_VACIA = "txml_cita_vacia"  # TX_ML
IND_TXCURR_COHORTE = "txcurr_cohorte_diff"  # Conciliación trimestral TX_CURR

DISPLAY_NAMES = {
    IND_NUM_GT_DEN:      "TX_PVLS (Num) > TX_PVLS (Den)",
    IND_DEN_GT_CURR:     "TX_PVLS (Den) > TX_CURR",
    IND_CD4_MISSING:     "CD4 vacío positivo",
    IND_TARV_GT_DIAG:    "Fecha inicio TARV < Fecha diagnóstico (HTS_TST)",
    IND_DIAG_BAD_FMT:    "Formato fecha diagnóstico",
    IND_CURR_Q1Q2_DIFF:  "TX_CURR ≠ Dispensación_TARV",
    IND_ID_DUPLICADO:    "ID (expediente) duplicado",
    IND_SEXO_INVALID:    "Sexo inválido (HTS_TST)",
}
DISPLAY_NAMES.update({
    IND_TXML_CITA_VACIA: "TX_ML: Última cita esperada vacía",  # TX_ML
    IND_TXCURR_COHORTE: "Conciliación TX_CURR trimestral",
})

MESES = {
    "enero","febrero","marzo","abril","mayo","junio",
    "julio","agosto","septiembre","octubre","noviembre","diciembre",
    "ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic",
    "january","february","march","april","may","june",
    "july","august","september","october","november","december",
}
RUIDO_DIRS = {"", ".", "..", "__MACOSX", ".ds_store", ".git"}
INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'
SPAN_ABBR = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]

def _pct(e, c): return round((e / c * 100.0), 2) if c else 0.0

def _safe_sheet_name(name: str, used: set) -> str:
    base = unicodedata.normalize("NFKD", name).encode("ascii","ignore").decode("ascii")
    base = re.sub(INVALID_SHEET_CHARS, "-", base).strip() or "Hoja"
    base = base[:31]
    candidate = base; i = 1
    while candidate in used:
        suf = f"_{i}"; candidate = base[:31-len(suf)] + suf; i += 1
    used.add(candidate)
    return candidate

def _norm(s) -> str:
    if s is None: return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode("ascii")
    return re.sub(r"\s+", " ", s.strip().lower())

def _normalize_sexo(x) -> str:
    sx = _norm(x)
    if sx.startswith("masc"): return "Masculino"
    if sx.startswith("fem"):  return "Femenino"
    return str(x).strip()

def _is_nonempty_site(x) -> bool:
    """
    Devuelve True si 'Nombre del sitio' tiene información útil.
    Considera vacíos valores como '', 'NA', 'N/A', 'Desconocido', '-', '—', '.', etc.
    """
    if x is None:
        return False
    s = str(x).strip()
    if not s:
        return False
    # normalizado para comparar placeholders
    sn = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").strip().lower()
    invalid_tokens = {
        "na", "n/a", "desconocido", "sin sitio", "sin nombre", "no aplica", "no aplica.", "no-aplica",
        "nan", "none", "null", "s/n", "s/d", "s.i.", "sd"
    }
    if sn in invalid_tokens:
        return False
    # solo símbolos / separadores
    if re.fullmatch(r"[-_\.—–]+", s):
        return False
    return True




def buscar_columna_multi(columnas, *patrones) -> Optional[str]:
    cols_norm = [_norm(c) for c in columnas]
    for i, cn in enumerate(cols_norm):
        if all(p in cn for p in patrones):
            return columnas[i]
    return None

def month_label_from_value(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        dt = pd.to_datetime(v, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return f"{SPAN_ABBR[dt.month-1]} {dt.year}"
    except Exception:
        pass
    return str(v).strip()

SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}

def _parse_any_date(v) -> pd.Timestamp:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return pd.NaT
    if isinstance(v, (pd.Timestamp, datetime)):
        return pd.Timestamp(v)
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        try:
            return pd.Timestamp(datetime(1899, 12, 30) + timedelta(days=float(v)))
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return pd.NaT
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt
    except Exception:
        pass
    sn = _norm(s)
    m = re.search(r"(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})", sn)
    if m:
        dia = int(m.group(1))
        mes = SPANISH_MONTHS.get(m.group(2))
        anio = int(m.group(3))
        if mes:
            try:
                return pd.Timestamp(year=anio, month=mes, day=dia)
            except Exception:
                return pd.NaT
    return pd.NaT

def _qfy_from_date(dt: pd.Timestamp) -> Optional[str]:
    if pd.isna(dt):
        return None
    mes = int(dt.month)
    anio = int(dt.year)
    fy = anio + 1 if mes >= 10 else anio
    if mes in (10, 11, 12):
        q = 1
    elif mes in (1, 2, 3):
        q = 2
    elif mes in (4, 5, 6):
        q = 3
    else:
        q = 4
    return f"Q{q} FY{str(fy)[-2:]}"

def _normalize_qfy(raw_qfy, fallback_date=None) -> Optional[str]:
    if raw_qfy is not None and not (isinstance(raw_qfy, float) and pd.isna(raw_qfy)):
        s = str(raw_qfy).strip()
        m = re.search(r"Q\s*([1-4])\s*FY\s*(\d{2,4})", s, re.IGNORECASE)
        if m:
            fy = int(m.group(2))
            fy = fy + 2000 if fy < 100 else fy
            return f"Q{int(m.group(1))} FY{str(fy)[-2:]}"
    dt = _parse_any_date(fallback_date if fallback_date is not None else raw_qfy)
    return _qfy_from_date(dt)

def _prev_qfy(qfy: str) -> Optional[str]:
    m = re.search(r"Q\s*([1-4])\s*FY\s*(\d{2,4})", str(qfy).strip(), re.IGNORECASE)
    if not m:
        return None
    q = int(m.group(1))
    fy = int(m.group(2))
    fy = fy + 2000 if fy < 100 else fy
    if q > 1:
        return f"Q{q-1} FY{str(fy)[-2:]}"
    return f"Q4 FY{str(fy-1)[-2:]}"

def _qfy_sort_key(qfy: str):
    m = re.search(r"Q\s*([1-4])\s*FY\s*(\d{2,4})", str(qfy).strip(), re.IGNORECASE)
    if not m:
        return (9999, 9)
    fy = int(m.group(2))
    fy = fy + 2000 if fy < 100 else fy
    return (fy, int(m.group(1)))


def _canon_text(v) -> str:
    s = _norm(v).replace("_", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _infer_context_date_from_text(text_value) -> pd.Timestamp:
    if text_value is None:
        return pd.NaT
    s = _canon_text(text_value)
    if not s:
        return pd.NaT

    month_num = None
    for token in re.split(r"[^a-z0-9]+", s):
        if token in SPANISH_MONTHS:
            month_num = SPANISH_MONTHS[token]
            break

    fy_year = None
    m_fy = re.search(r"fy\s*(\d{2,4})", s, re.IGNORECASE)
    if m_fy:
        fy_year = int(m_fy.group(1))
        fy_year = fy_year + 2000 if fy_year < 100 else fy_year

    calendar_year = None
    m_y4 = re.search(r"(20\d{2})", s)
    if m_y4:
        calendar_year = int(m_y4.group(1))
    elif month_num is not None:
        m_y2 = re.search(r"(\d{2})(?!\d)", s)
        if m_y2:
            calendar_year = 2000 + int(m_y2.group(1))

    if month_num is None:
        return pd.NaT

    if fy_year is not None:
        year = fy_year - 1 if month_num >= 10 else fy_year
    elif calendar_year is not None:
        year = calendar_year
    else:
        return pd.NaT

    try:
        return pd.Timestamp(year=year, month=month_num, day=1) + pd.offsets.MonthEnd(0)
    except Exception:
        return pd.NaT

def _qfy_from_context_text(*texts) -> Tuple[Optional[str], pd.Timestamp]:
    for t in texts:
        dt = _infer_context_date_from_text(t)
        if pd.notna(dt):
            return _qfy_from_date(dt), dt
    return None, pd.NaT

def _join_unique_text(values, sep=" | ", max_items=6) -> str:
    seen = []
    for v in values:
        s = str(v).strip()
        if not s or s.lower() in {"nan", "none"}:
            continue
        if s not in seen:
            seen.append(s)
    if len(seen) <= max_items:
        return sep.join(seen)
    return sep.join(seen[:max_items]) + f" | (+{len(seen)-max_items} más)"

def _age_columns_tx_curr(columns: List[str]) -> List[str]:
    out = []
    for c in columns:
        cn = _norm(c)
        if not cn or "rango" in cn:
            continue
        if re.fullmatch(r"<\s*1 ano", cn) or re.fullmatch(r"\d{1,2}\s*-\s*\d{1,2}\s*anos?", cn) or cn.startswith("65"):
            out.append(c)
    return out

def inferir_pais_mes(path_rel: str, default_pais: str, default_mes: str):
    ruta = path_rel.replace("\\", "/")
    partes = [p for p in ruta.split("/") if p.strip().lower() not in RUIDO_DIRS]
    if partes and partes[-1].lower().endswith(".xlsx"): partes = partes[:-1]
    pais = partes[-2].strip() if len(partes) >= 2 else default_pais
    # Evita meses como país
    if any(tok in MESES for tok in re.split(r"[_\-\s/\.]+", _norm(pais))):
        pais = default_pais
    pais = pais or default_pais
    mes = None
    for seg in reversed(partes):
        toks = re.split(r"[_\-\s/\.]+", _norm(seg))
        if any(t in MESES for t in toks):
            mes = seg.strip(); break
    if not mes:
        base = os.path.basename(path_rel)
        base = re.sub(r"\.xlsx$", "", base, flags=re.IGNORECASE)
        toks = re.split(r"[_\-\s/\.]+", _norm(base))
        if any(t in MESES for t in toks):
            mes = base
    mes = mes or default_mes
    return pais, mes

def leer_excel_desde_bytes(nombre, data_bytes): return pd.ExcelFile(io.BytesIO(data_bytes))

def encontrar_fila_encabezado(df_raw: pd.DataFrame, needles) -> Optional[int]:
    try:
        mask = df_raw.astype(str).apply(
            lambda r: all(any(needle.lower() in str(x).lower() for x in r.values) for needle in needles), axis=1
        ); idxs = df_raw[mask].index.tolist()
        return idxs[0] if idxs else None
    except Exception:
        return None

def normalizar_tabla_por_encabezado(df_raw: pd.DataFrame, idx_header: int):
    columnas = df_raw.iloc[idx_header].fillna("").astype(str).tolist()
    df_body = df_raw.iloc[idx_header + 1:].copy(); df_body.columns = columnas
    return df_body.reset_index(drop=True), columnas

def numeros_seguro(v): return pd.to_numeric(v, errors="coerce")

def _add_metric(ind_key: str, pais: str, mes_rep: str, depto: str = "", sitio: str = "",
                checks_add: int = 0, errors_add: int = 0):
    if checks_add:
        st.session_state.metrics_global[ind_key]["checks"] += checks_add
        st.session_state.metrics_by_pds[(pais or "", depto or "", sitio or "", mes_rep or "", ind_key)]["checks"] += checks_add
    if errors_add:
        st.session_state.metrics_global[ind_key]["errors"] += errors_add
        st.session_state.metrics_by_pds[(pais or "", depto or "", sitio or "", mes_rep or "", ind_key)]["errors"] += errors_add

def _rename_standard_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    for c in df.columns:
        cn = _norm(c)
        if not cn:
            continue
        if "sexo" in cn or "genero" in cn or "género" in cn:
            mapping[c] = "Sexo"
        elif ("tipo" in cn and "pobl" in cn) or "poblacion clave" in cn or "población clave" in cn:
            mapping[c] = "Tipo de población"
        elif "pais" in cn:
            mapping[c] = "País"
        elif "departamento" in cn or "depto" in cn or "provincia" in cn:
            mapping[c] = "Departamento"
        elif ("servicio" in cn and "salud" in cn) or "sitio" in cn or "clinica" in cn or "clínica" in cn:
            mapping[c] = "Sitio"
        elif "mes" in cn and "rep" in cn:
            mapping[c] = "Mes de reporte"
        elif "fecha" in cn and "reporte" in cn:
            mapping[c] = "Fecha de reporte"
    return df.rename(columns=mapping)

def _dedupe_columns(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}__{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out

def _coerce_scalar(v):
    if isinstance(v, pd.Series):
        for x in v.tolist():
            if pd.notna(x) and str(x).strip():
                return x
        return ""
    return v

def _first_col(df: pd.DataFrame, *tokens) -> Optional[str]:
    for c in df.columns:
        cn = _norm(c)
        if all(t in cn for t in tokens):
            return c
    return None

def show_df_or_note(df, note="— Sin filas para mostrar —", height=300):
    if df is None or (isinstance(df, pd.DataFrame) and df.empty):
        st.caption(note); return False
    st.dataframe(df, use_container_width=True, height=height); return True

# ============================
# ------- VALIDACIONES -------
# ============================
def procesar_tx_pvls_y_curr(
    xl: pd.ExcelFile, pais_inferido: str, mes_inferido: str, nombre_archivo: str,
    errores_numerador, errores_txpvls
):
    if "TX_PVLS" not in xl.sheet_names: return
    pvls_raw = xl.parse("TX_PVLS", header=None)
    idx_header = encontrar_fila_encabezado(pvls_raw, ["Sexo", "Tipo"])
    if idx_header is None: return
    df_data, columnas = normalizar_tabla_por_encabezado(pvls_raw, idx_header)
    df_data = _rename_standard_columns(df_data)

    # Localizar filas de numerador/denominador
    try:
        idx_num = df_data[df_data.iloc[:,0].astype(str).str.contains("Numerador", case=False, na=False)].index[0]
        idx_den = df_data[df_data.iloc[:,0].astype(str).str.contains("Denominador", case=False, na=False)].index[0]
    except IndexError:
        try:
            idx_num = df_data[df_data.get("TX_PVLS Numerador", "").astype(str).str.contains("Numerador", case=False, na=False)].index[0]
            idx_den = df_data[df_data.get("TX_PVLS Numerador", "").astype(str).str.contains("Denominador", case=False, na=False)].index[0]
        except Exception:
            return

    df_num = _rename_standard_columns(df_data.iloc[idx_num + 1:idx_den].copy())
    df_den = _rename_standard_columns(df_data.iloc[idx_den + 1:].copy())

    # Columnas de edad (tolerante)
    columnas_edad = [c for c in df_data.columns
                     if ("año" in c.lower()) or ("ano" in _norm(c)) or ("65" in c) or ("+" in c) or ("más" in c.lower() and "65" in c.lower())]

    # País/Depto/Sitio
    col_pais   = buscar_columna_multi(df_data.columns, "pais")
    col_depto  = buscar_columna_multi(df_data.columns, "departamento") or buscar_columna_multi(df_data.columns, "depto") or buscar_columna_multi(df_data.columns, "provincia")
    col_sitio  = buscar_columna_multi(df_data.columns, "servicio", "salud") or buscar_columna_multi(df_data.columns, "sitio") or buscar_columna_multi(df_data.columns, "clinica")

    # Fecha/Mes de reporte (prioridad)
    col_fecha_rep = buscar_columna_multi(df_data.columns, "fecha", "reporte")
    col_mesrep    = buscar_columna_multi(df_data.columns, "mes", "reporte")
    def _ctx(row):
        p = str(row.get(col_pais)) if col_pais else pais_inferido
        d = str(row.get(col_depto)) if col_depto else ""
        s = str(row.get(col_sitio)) if col_sitio else ""
        raw_mes = row.get(col_fecha_rep) if col_fecha_rep else (row.get(col_mesrep) if col_mesrep else None)
        m = month_label_from_value(raw_mes) or month_label_from_value(mes_inferido)
        p = p if str(p).strip() else pais_inferido
        d = d if str(d).strip() else ""
        s = s if str(s).strip() else ""
        m = m if str(m).strip() else month_label_from_value(mes_inferido)
        return p, d, s, m

    fila_base_num = idx_header + 3 + idx_num + 1

    # Numerador > Denominador
    for i, row_num in df_num.iterrows():
        sexo = str(row_num.get("Sexo", "")).strip()
        pob  = str(row_num.get("Tipo de población", "")).strip()
        if _normalize_sexo(sexo) not in ["Masculino", "Femenino"]: continue
        row_den = df_den[(df_den["Sexo"].astype(str).str.replace("_", " ", regex=False).str.strip()==sexo) &
                         (df_den["Tipo de población"].astype(str).str.replace("_", " ", regex=False).str.strip()==pob)]
        if row_den.empty: continue
        row_den = row_den.iloc[0]
        pais_row, depto_row, sitio_row, mes_rep = _ctx(row_num)

        for col in columnas_edad:
            val_num = numeros_seguro(row_num.get(col))
            val_den = numeros_seguro(row_den.get(col))
            if pd.notna(val_num) and pd.notna(val_den):
                _add_metric(IND_NUM_GT_DEN, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
                if val_num > val_den:
                    _add_metric(IND_NUM_GT_DEN, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                    col_idx = df_data.columns.tolist().index(col)
                    errores_numerador.append({
                        "País": pais_row, "Departamento": depto_row, "Sitio": sitio_row, "Mes de reporte": mes_rep,
                        "Archivo": nombre_archivo, "Sexo": sexo, "Tipo de población": pob, "Rango de edad": col,
                        "Numerador": float(val_num), "Denominador": float(val_den),
                        "Fila Excel": int(fila_base_num + i), "Columna Excel": get_column_letter(col_idx + 1)
                    })

    # Denominador (PVLS) > TX_CURR
    if "TX_CURR" in xl.sheet_names:
        curr_raw = xl.parse("TX_CURR", header=None)
        idx_curr = encontrar_fila_encabezado(curr_raw, ["Sexo", "Tipo"])
        if idx_curr is None: return
        df_curr, _ = normalizar_tabla_por_encabezado(curr_raw, idx_curr)
        df_curr = _rename_standard_columns(df_curr)
        fila_base_excel_den = idx_header + 3 + idx_den + 1

        for i, row_den in df_den.iterrows():
            sexo = str(row_den.get("Sexo", "")).strip()
            pob  = str(row_den.get("Tipo de población", "")).strip()
            if _normalize_sexo(sexo) not in ["Masculino", "Femenino"]: continue
            row_curr = df_curr[(df_curr["Sexo"].astype(str).str.replace("_", " ", regex=False).str.strip()==sexo) &
                               (df_curr["Tipo de población"].astype(str).str.replace("_", " ", regex=False).str.strip()==pob)]
            if row_curr.empty: continue
            row_curr = row_curr.iloc[0]

            pais_row, depto_row, sitio_row, mes_rep = _ctx(row_den)
            for col in columnas_edad:
                val_den = numeros_seguro(row_den.get(col))
                val_curr = numeros_seguro(row_curr.get(col))
                if pd.notna(val_den) and pd.notna(val_curr):
                    _add_metric(IND_DEN_GT_CURR, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
                    if val_den > val_curr:
                        _add_metric(IND_DEN_GT_CURR, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                        col_idx = df_data.columns.tolist().index(col)
                        errores_txpvls.append({
                            "País": pais_row, "Departamento": depto_row, "Sitio": sitio_row, "Mes de reporte": mes_rep,
                            "Archivo": nombre_archivo, "Sexo": sexo, "Tipo de población": pob, "Rango de edad": col,
                            "Denominador (PVLS)": float(val_den), "TX_CURR": float(val_curr),
                            "Fila Excel": int(fila_base_excel_den + i), "Columna Excel": get_column_letter(col_idx + 1)
                        })

# ===== TX_CURR vs Dispensación_TARV (por Sexo/Edad) =====
def procesar_tx_curr_cuadros(
    xl: pd.ExcelFile, pais_inferido: str, mes_inferido: str,
    nombre_archivo: str, errores_currq
):
    sheet_name = "TX_CURR"
    if sheet_name not in xl.sheet_names: return
    df_raw = xl.parse(sheet_name, header=None)
    nrows, ncols = df_raw.shape

    def _find_label_positions(alternatives: List[List[str]]) -> List[Tuple[int,int]]:
        pos = []
        for r in range(nrows):
            for c in range(ncols):
                s = _norm(df_raw.iat[r, c])
                if not s: continue
                for toks in alternatives:
                    if all(tok in s for tok in toks):
                        pos.append((r, c)); break
        return pos

    pos_tx = _find_label_positions([["tx", "curr"]])
    pos_et = _find_label_positions([
        ["dispens","tar"], ["dispensacion","tar"], ["dispensación","tar"],
        ["entrega","tar"], ["entrega","tavr"]
    ])
    if not pos_tx or not pos_et:
        return

    fila_tx = min(pos_tx)[0]
    fila_et = min(pos_et)[0]

    def _find_header_after(start_row: int) -> Optional[int]:
        for r in range(start_row, min(start_row + 50, nrows)):
            row_vals = df_raw.iloc[r].tolist()
            if any("sexo" in _norm(x) for x in row_vals):
                cols = [str(x) for x in row_vals]
                cols_norm = [_norm(x) for x in cols]
                try:
                    col_sexo = next(i for i, cn in enumerate(cols_norm) if "sexo" in cn)
                except StopIteration:
                    continue
                edad_ok = any(
                    ("ano" in _norm(cols[j])) or ("año" in _norm(cols[j])) or re.search(r"\b65\b", _norm(cols[j])) or
                    ("+" in _norm(cols[j])) or ("mas" in _norm(cols[j]) and "65" in _norm(cols[j]))
                    for j in range(col_sexo + 1, ncols)
                )
                if edad_ok:
                    return r
        return None

    hdr_tx = _find_header_after(fila_tx)
    hdr_et = _find_header_after(fila_et)
    if hdr_tx is None or hdr_et is None: return

    def _extract_table_totals(header_row: int, stop_row: Optional[int]):
        cols = df_raw.iloc[header_row].fillna("").astype(str).tolist()
        cols_norm = [_norm(x) for x in cols]
        try:
            col_sexo = next(i for i, cn in enumerate(cols_norm) if "sexo" in cn)
        except StopIteration:
            return {}, {}, None

        edad_idx, edad_key, edad_map = [], [], {}
        for j in range(col_sexo + 1, ncols):
            lab  = cols[j]; labn = _norm(lab)
            if ("ano" in labn) or ("año" in labn) or re.search(r"\b65\b", labn) or ("+" in labn) or ("mas" in labn and "65" in labn):
                k = labn
                edad_idx.append(j); edad_key.append(k); edad_map[k] = str(lab)

        if not edad_idx:
            return {}, {}, col_sexo

        totals: Dict[Tuple[str,str], int] = {}
        r = header_row + 1
        while r < nrows:
            if stop_row is not None and r >= stop_row: break
            rown = [_norm(x) for x in df_raw.iloc[r].tolist()]
            if any("sexo" in x for x in rown): break
            sx_raw = df_raw.iat[r, col_sexo]
            sx = _normalize_sexo(sx_raw)
            if sx in ["Masculino", "Femenino"]:
                for c_idx, ekey in zip(edad_idx, edad_key):
                    v = pd.to_numeric(df_raw.iat[r, c_idx], errors="coerce")
                    if pd.notna(v):
                        v = int(round(float(v)))
                        key = (sx, ekey)
                        totals[key] = totals.get(key, 0) + v
            r += 1

        return totals, edad_map, col_sexo

    if hdr_tx < hdr_et:
        totals_tx, edades_tx, _ = _extract_table_totals(hdr_tx, hdr_et)
        totals_et, edades_et, _ = _extract_table_totals(hdr_et, None)
    else:
        totals_et, edades_et, _ = _extract_table_totals(hdr_et, hdr_tx)
        totals_tx, edades_tx, _ = _extract_table_totals(hdr_tx, None)

    # Contexto
    cols_hdr = df_raw.iloc[hdr_tx].fillna("").astype(str).tolist()
    cols_hdrn = [_norm(x) for x in cols_hdr]

    def _find_col_idx(cols_norm, *patrones):
        for i, cn in enumerate(cols_norm):
            if all(p in cn for p in patrones):
                return i
        return None

    col_pais_i   = _find_col_idx(cols_hdrn, "pais")
    col_depto_i  = (_find_col_idx(cols_hdrn, "departamento") or
                    _find_col_idx(cols_hdrn, "depto") or
                    _find_col_idx(cols_hdrn, "provincia"))
    col_sitio_i  = (_find_col_idx(cols_hdrn, "servicio") or
                    _find_col_idx(cols_hdrn, "salud") or
                    _find_col_idx(cols_hdrn, "sitio") or
                    _find_col_idx(cols_hdrn, "clinica"))
    col_mesrep_i = (_find_col_idx(cols_hdrn, "fecha") if _find_col_idx(cols_hdrn, "reporte") is not None else None)
    if col_mesrep_i is None:
        col_mesrep_i = _find_col_idx(cols_hdrn, "mes")

    def _ctx_from_rowvals(row_vals):
        def _val(idx, fallback):
            try:
                v = row_vals[idx] if idx is not None else fallback
            except Exception:
                v = fallback
            v = str(v).strip()
            return v if v else fallback
        p = _val(col_pais_i, pais_inferido)
        d = _val(col_depto_i, "")
        s = _val(col_sitio_i, "")
        raw_mes = row_vals[col_mesrep_i] if col_mesrep_i is not None else None
        m = month_label_from_value(raw_mes) or month_label_from_value(mes_inferido)
        return p, d, s, m

    fila_ctx_vals = df_raw.iloc[hdr_tx + 1].fillna("").astype(str).tolist() if (hdr_tx + 1) < nrows else []
    pais_row, depto_row, sitio_row, mes_rep = _ctx_from_rowvals(fila_ctx_vals)

    all_keys = set(totals_tx.keys()) | set(totals_et.keys())
    for (sexo, edad_key) in sorted(all_keys):
        v_tx = int(totals_tx.get((sexo, edad_key), 0))
        v_et = int(totals_et.get((sexo, edad_key), 0))
        etiqueta_edad = (edades_tx.get(edad_key) or edades_et.get(edad_key) or edad_key)
        _add_metric(IND_CURR_Q1Q2_DIFF, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
        if v_tx != v_et:
            _add_metric(IND_CURR_Q1Q2_DIFF, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
            errores_currq.append({
                "País": pais_row, "Departamento": depto_row, "Sitio": sitio_row, "Mes de reporte": mes_rep,
                "Archivo": nombre_archivo, "Sexo": sexo, "Rango de edad": etiqueta_edad,
                "TX_CURR": v_tx, "Dispensación_TARV": v_et,
                "Diferencia (TX_CURR - Disp_TARV)": v_tx - v_et,
                "Disp_TARV > TX_CURR": "Sí" if v_et > v_tx else "No",
            })

# ===== HTS_TST + ID Duplicado + Sexo inválido + CD4 vacío positivo =====
def procesar_hts_tst(
    xl: pd.ExcelFile, pais_inferido: str, mes_inferido: str, nombre_archivo: str,
    errores_cd4, errores_formato_fecha_diag, errores_iddup,
    errores_sexo, errores_tarv_gt
):
    if "HTS_TST" not in xl.sheet_names:
        return

    df_raw = xl.parse("HTS_TST", header=None)
    idx_hts = encontrar_fila_encabezado(df_raw, ["Resultado", "CD4"])
    if idx_hts is None:
        return

    df_data, _ = normalizar_tabla_por_encabezado(df_raw, idx_hts)
    df_data.columns = _dedupe_columns(df_data.columns)
    df_data = _rename_standard_columns(df_data)

    # Columnas base
    col_resultado = _first_col(df_data, "resultado")  # genérico
    col_resultado_vih = (
        _first_col(df_data, "resultado", "vih") or
        _first_col(df_data, "resultado de la prueba de vih") or
        col_resultado
    )
    col_cd4       = _first_col(df_data, "cd4")

    # Búsqueda por nombre normalizado exacto primero, luego patrones parciales
    _cols_norm_map = {_norm(c): c for c in df_data.columns}
    col_tarv      = (
        _cols_norm_map.get("fecha de inicio tarv") or
        _cols_norm_map.get("fecha inicio tarv") or
        _cols_norm_map.get("fecha de inicio de tarv") or
        _first_col(df_data, "inicio", "tarv") or
        _first_col(df_data, "inicio", "tar") or
        _first_col(df_data, "fecha", "tarv") or
        _first_col(df_data, "tarv")
    )
    col_diag      = (
        _cols_norm_map.get("fecha del diagnostico") or
        _cols_norm_map.get("fecha de diagnostico") or
        _cols_norm_map.get("fecha diagnostico") or
        _first_col(df_data, "fecha", "diagnostico") or
        _first_col(df_data, "fecha", "diagn") or
        _first_col(df_data, "diagnostico")
    )
    col_sitio     = _first_col(df_data, "servicio", "salud") or _first_col(df_data, "sitio") or _first_col(df_data, "clinica")
    col_pais      = _first_col(df_data, "pais")
    col_depto     = _first_col(df_data, "departamento") or _first_col(df_data, "depto") or _first_col(df_data, "provincia")
    col_id        = (_first_col(df_data, "id", "expediente") or
                     _first_col(df_data, "numero", "expediente") or
                     _first_col(df_data, "número", "expediente") or
                     _first_col(df_data, "id"))
    col_sexo      = _first_col(df_data, "sexo") or _first_col(df_data, "genero") or _first_col(df_data, "género")
    col_edad      = _first_col(df_data, "edad")
    col_motivo    = _first_col(df_data, "motivo", "cd4")

    # Diagnóstico: excluir columnas que también contengan "tarv"/"tar" para no confundir con TARV
    if col_diag and col_tarv and col_diag == col_tarv:
        col_diag = None

    if not all([col_resultado, col_cd4, col_diag]):
        st.warning(
            f"⚠️ **{nombre_archivo}** – HTS_TST: no se encontraron columnas clave. "
            f"Resultado=`{col_resultado}` | CD4=`{col_cd4}` | Diagnóstico=`{col_diag}`. "
            f"Columnas detectadas: {list(df_data.columns)[:15]}"
        )
        return

    if col_tarv is None:
        st.warning(
            f"⚠️ **{nombre_archivo}** – HTS_TST: no se encontró la columna 'Fecha de inicio TARV'. "
            f"La validación de fechas TARV vs diagnóstico no se ejecutará. "
            f"Columnas detectadas: {list(df_data.columns)[:15]}"
        )

    # Helpers de fecha
    def _excel_serial_to_datetime(n):
        base = datetime(1899, 12, 30)  # Excel bug 1900
        days = int(float(n))
        frac = float(n) - days
        return base + timedelta(days=days) + timedelta(days=frac)

    def es_fecha_valida(fecha_val) -> bool:
        # NaN / vacío no se valida aquí (se hace afuera via tiene_valor)
        if isinstance(fecha_val, (pd.Timestamp, datetime)):
            return True
        if isinstance(fecha_val, (int, float)) and not (isinstance(fecha_val, float) and math.isnan(fecha_val)):
            try:
                _excel_serial_to_datetime(fecha_val)
                return True
            except Exception:
                return False
        s = str(fecha_val).strip()
        if not s:
            return False
        s_norm = s.replace(".", "/").replace("-", "/")
        try:
            pd.to_datetime(s_norm, dayfirst=True, errors="raise")
            return True
        except Exception:
            return False

    fila_base_hts = idx_hts + 2
    for i, row in df_data.iterrows():
        resultado  = str(_coerce_scalar(row.get(col_resultado))).strip().lower()
        cd4        = _coerce_scalar(row.get(col_cd4))
        fecha_diag = _coerce_scalar(row.get(col_diag))
        fecha_tarv = _coerce_scalar(row.get(col_tarv)) if col_tarv else None
        sitio      = _coerce_scalar(row.get(col_sitio)) if col_sitio else ""
        pais_row   = _coerce_scalar(row.get(col_pais))  if col_pais else pais_inferido
        depto_row  = _coerce_scalar(row.get(col_depto)) if col_depto else ""

        mes_rep    = month_label_from_value(fecha_diag) or month_label_from_value(mes_inferido)

        pais_row   = str(pais_row).strip() or pais_inferido
        depto_row  = str(depto_row).strip()
        sitio_row  = str(sitio).strip()
        mes_rep    = str(mes_rep).strip() or month_label_from_value(mes_inferido)

        # === Validación de SEXO permitido (cuenta chequeo si hay algo en celda) ===
        if col_sexo is not None:
            sx_raw = _coerce_scalar(row.get(col_sexo))
            if not (pd.isna(sx_raw) or str(sx_raw).strip() == ""):
                _add_metric(IND_SEXO_INVALID, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
                sx_norm = _normalize_sexo(sx_raw)
                if sx_norm not in ["Masculino", "Femenino"]:
                    _add_metric(IND_SEXO_INVALID, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                    errores_sexo.append({
                        "País": pais_row,
                        "Departamento": depto_row,
                        "Sitio": sitio_row,
                        "Mes de reporte": mes_rep,
                        "Archivo": nombre_archivo,
                        "Sexo (valor encontrado)": str(sx_raw).strip(),
                        "Sugerido (normalizado)": sx_norm if sx_norm in ["Masculino","Femenino"] else "",
                        "Fila Excel": int(fila_base_hts + i),
                        "Columna Excel": col_sexo
                    })

        # CHECK duplicados (para %)
        if col_id is not None:
            id_val = _coerce_scalar(row.get(col_id))
            if pd.notna(id_val) and str(id_val).strip():
                _add_metric(IND_ID_DUPLICADO, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)

        # CD4 vacío cuando Resultado = Positivo (agrega Motivo de no CD4 después de Resultado)
        if resultado == "positivo":
            _add_metric(IND_CD4_MISSING, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
            if pd.isna(cd4) or str(cd4).strip() == "":
                _add_metric(IND_CD4_MISSING, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                errores_cd4.append({
                    "País": pais_row,
                    "Departamento": depto_row,
                    "Sitio": sitio_row,
                    "Mes de reporte": mes_rep,
                    "Archivo": nombre_archivo,
                    "Resultado": "Positivo",
                    "Motivo de no CD4": ("" if col_motivo is None else str(_coerce_scalar(row.get(col_motivo))).strip()),
                    "CD4 Basal": "",
                    "Fila Excel": int(fila_base_hts + i),
                    "Columna Excel": col_cd4
                })

        # Fecha inicio TARV no debe ser menor que Fecha del diagnóstico
        if col_tarv and pd.notna(fecha_diag) and pd.notna(fecha_tarv) and str(fecha_tarv).strip() not in ("", "nan", "None"):
            try:
                def _parse_fecha(v):
                    if isinstance(v, (pd.Timestamp, datetime)):
                        return pd.Timestamp(v)
                    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
                        # Serial de Excel → datetime
                        base = datetime(1899, 12, 30)
                        return pd.Timestamp(base + timedelta(days=float(v)))
                    s = str(v).strip().replace(".", "/").replace("-", "/")
                    return pd.to_datetime(s, dayfirst=True, errors="coerce")

                fd = _parse_fecha(fecha_diag)
                ft = _parse_fecha(fecha_tarv)
                if pd.notna(fd) and pd.notna(ft):
                    _add_metric(IND_TARV_GT_DIAG, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
                    if ft < fd:
                        _add_metric(IND_TARV_GT_DIAG, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                        col_diag_idx = list(df_data.columns).index(col_diag) if col_diag in df_data.columns else None
                        col_tarv_idx = list(df_data.columns).index(col_tarv) if col_tarv and col_tarv in df_data.columns else None
                        errores_tarv_gt.append({
                            "País": pais_row,
                            "Departamento": depto_row,
                            "Sitio": sitio_row,
                            "ID expediente": str(_coerce_scalar(row.get(col_id))).strip() if col_id else "",
                            "Mes de reporte": mes_rep,
                            "Archivo": nombre_archivo,
                            "Fecha diagnóstico": fd.date(),
                            "Fecha inicio TARV": ft.date(),
                            "Fila Excel": int(fila_base_hts + i),
                            "Columna Excel Diagnóstico": get_column_letter(col_diag_idx + 1) if col_diag_idx is not None else col_diag,
                            "Columna Excel TARV": get_column_letter(col_tarv_idx + 1) if col_tarv_idx is not None else col_tarv,
                        })
            except Exception as _ex_tarv:
                st.warning(f"⚠️ Error comparando fechas TARV fila {fila_base_hts+i} en {nombre_archivo}: {_ex_tarv}")

        # Formato de Fecha Diagnóstico (si celda tiene valor)
        tiene_valor = not (pd.isna(fecha_diag) or str(fecha_diag).strip() == "")
        if tiene_valor:
            _add_metric(IND_DIAG_BAD_FMT, pais_row, mes_rep, depto_row, sitio_row, checks_add=1)
            if not es_fecha_valida(fecha_diag):
                _add_metric(IND_DIAG_BAD_FMT, pais_row, mes_rep, depto_row, sitio_row, errors_add=1)
                errores_formato_fecha_diag.append({
                    "País": pais_row,
                    "Departamento": depto_row,
                    "Sitio": sitio_row,
                    "Mes de reporte": mes_rep,
                    "Archivo": nombre_archivo,
                    "Fecha del diagnóstico de la prueba": fecha_diag,
                    "Fila Excel": int(fila_base_hts + i),
                    "Columna Excel": col_diag
                })

    # Duplicados ID (una fila POR REGISTRO duplicado con Resultado VIH)
    if col_id:
        try:
            col_id_idx = list(df_data.columns).index(col_id)
        except ValueError:
            col_id_idx = None

        ids_raw = df_data[col_id].astype(str).str.replace("_", " ", regex=False).str.strip()
        mask_non_empty = ids_raw.replace({"nan": "", "NaN": ""}).astype(bool)
        vc = ids_raw[mask_non_empty].value_counts()
        duplicados = vc[vc > 1]

        if not duplicados.empty:
            for id_val, count in duplicados.items():
                idxs = df_data.index[ids_raw == id_val].tolist()
                r_ref = df_data.loc[idxs[0]]
                ref_pais  = str(_coerce_scalar(r_ref.get(col_pais)))  if col_pais  else pais_inferido
                ref_depto = str(_coerce_scalar(r_ref.get(col_depto))) if col_depto else ""
                ref_sitio = str(_coerce_scalar(r_ref.get(col_sitio)))  if col_sitio  else ""
                ref_mes   = month_label_from_value(_coerce_scalar(r_ref.get(col_diag))) or month_label_from_value(mes_inferido)
                _add_metric(IND_ID_DUPLICADO,
                            ref_pais.strip() or pais_inferido,
                            ref_mes.strip() or month_label_from_value(mes_inferido),
                            ref_depto.strip(),
                            ref_sitio.strip(),
                            errors_add=int(count) - 1)

                for i in idxs:
                    r = df_data.loc[i]
                    pais_row  = str(_coerce_scalar(r.get(col_pais)))  if col_pais  else pais_inferido
                    depto_row = str(_coerce_scalar(r.get(col_depto))) if col_depto else ""
                    sitio_row  = str(_coerce_scalar(r.get(col_sitio)))  if col_sitio  else ""
                    mes_rep   = month_label_from_value(_coerce_scalar(r.get(col_diag))) or month_label_from_value(mes_inferido)

                    pais_row  = pais_row.strip() or pais_inferido
                    depto_row = depto_row.strip()
                    sitio_row = sitio_row.strip()
                    mes_rep   = mes_rep.strip() or month_label_from_value(mes_inferido)

                    fila_excel = int(idx_hts + 2 + i)
                    col_letter = get_column_letter(col_id_idx + 1) if col_id_idx is not None else col_id
                    resultado_vih_val = _coerce_scalar(r.get(col_resultado_vih)) if col_resultado_vih else ""

                    # === NUEVO: capturar Sexo y Edad para la tabla de salida de duplicados ===
                    sexo_raw = _coerce_scalar(r.get(col_sexo)) if col_sexo else ""
                    sexo_out = _normalize_sexo(sexo_raw) if str(sexo_raw).strip() else ""
                    edad_raw = _coerce_scalar(r.get(col_edad)) if col_edad else ""
                    edad_num = pd.to_numeric(edad_raw, errors="coerce")
                    if pd.notna(edad_num):
                        edad_out = int(edad_num) if float(edad_num).is_integer() else float(edad_num)
                    else:
                        edad_out = str(edad_raw).strip()

                    errores_iddup.append({
                        "País": pais_row,
                        "Departamento": depto_row,
                        "Sitio": sitio_row,
                        "Mes de reporte": mes_rep,
                        "Archivo": nombre_archivo,
                        "ID expediente": str(id_val),
                        "Sexo": sexo_out,
                        "Edad": edad_out,
                        "Fecha del diagnóstico": (
                            (pd.to_datetime(_coerce_scalar(r.get(col_diag)), dayfirst=True, errors="coerce").date() if pd.notna(pd.to_datetime(_coerce_scalar(r.get(col_diag)), dayfirst=True, errors="coerce")) else str(_coerce_scalar(r.get(col_diag))).strip()) if col_diag else ""
                        ),
                        "Resultado prueba VIH": str(resultado_vih_val),
                        "Fila Excel": fila_excel,
                        "Columna Excel": col_letter,
                        "Ocurrencias ID": int(count),
                    })

# ===== VALIDACIÓN TX_ML: Fecha de su última cita esperada + Modalidad =====
def procesar_tx_ml_cita(
    xl: pd.ExcelFile, pais_inferido: str, mes_inferido: str, nombre_archivo: str,
    errores_txml_cita
):
    """
    TX_ML → Valida que 'Fecha de su última cita esperada' NO venga vacía.
    • Aplica FILTRO por 'Nombre del sitio' SOLO si la columna existe.
    • En la salida incluye 'Modalidad de reporte' justo después de 'ID expediente'.
    """
    sheet_name = "TX_ML"
    if sheet_name not in xl.sheet_names:
        return

    # Leer hoja en crudo y ubicar encabezado por tokens robustos
    df_raw = xl.parse(sheet_name, header=None)
    nrows, ncols = df_raw.shape

    def _row_has_tokens_norm(row_vals, tokens):
        for cell in row_vals:
            s = _norm(cell)
            if all(tok in s for tok in tokens):
                return True
        return False

    idx_header = None
    # Preferente: fecha + última + cita + esper(ada)
    for r in range(nrows):
        if _row_has_tokens_norm(df_raw.iloc[r].tolist(), ["fecha", "ultima", "cita", "esper"]):
            idx_header = r
            break
    # Fallback amplio: fecha + cita
    if idx_header is None:
        for r in range(nrows):
            if _row_has_tokens_norm(df_raw.iloc[r].tolist(), ["fecha", "cita"]):
                idx_header = r
                break
    if idx_header is None:
        return  # no se detectó cabecera

    # Normalizar tabla a partir del encabezado encontrado
    df_data, _ = normalizar_tabla_por_encabezado(df_raw, idx_header)
    df_data = _rename_standard_columns(df_data)

    # Columnas objetivo y de contexto (tolerantes a variantes)
    col_cita_esperada = (
        buscar_columna_multi(df_data.columns, "fecha", "ultima", "cita", "esper") or
        buscar_columna_multi(df_data.columns, "fecha", "cita", "esper") or
        buscar_columna_multi(df_data.columns, "fecha", "cita")
    )
    if not col_cita_esperada:
        return

    col_modalidad = (
        buscar_columna_multi(df_data.columns, "modalidad", "reporte") or
        buscar_columna_multi(df_data.columns, "modalidad")
    )
    col_pais      = buscar_columna_multi(df_data.columns, "pais")
    col_depto     = (buscar_columna_multi(df_data.columns, "departamento") or
                     buscar_columna_multi(df_data.columns, "depto") or buscar_columna_multi(df_data.columns, "provincia"))
    col_sitio     = (buscar_columna_multi(df_data.columns, "servicio", "salud") or
                     buscar_columna_multi(df_data.columns, "sitio") or
                     buscar_columna_multi(df_data.columns, "clinica"))
    col_fecha_rep = buscar_columna_multi(df_data.columns, "fecha", "reporte")
    col_mesrep    = buscar_columna_multi(df_data.columns, "mes", "reporte")
    col_id        = (buscar_columna_multi(df_data.columns, "id", "expediente") or
                     buscar_columna_multi(df_data.columns, "numero", "expediente") or
                     buscar_columna_multi(df_data.columns, "número", "expediente") or
                     buscar_columna_multi(df_data.columns, "id"))

    col_cita_exp_idx = df_data.columns.tolist().index(col_cita_esperada)
    fila_base_txml = idx_header + 2  # coherente con otras tablas

    for i, row in df_data.iterrows():
        # Señal mínima de fila (algo en ID/Sitio/País o en la propia columna objetivo)
        row_has_signal = any(
            str(_coerce_scalar(row.get(c))).strip()
            for c in [col_id, col_sitio, col_pais] if c is not None
        ) or not (pd.isna(row.get(col_cita_esperada)) or str(row.get(col_cita_esperada)).strip() == "")
        if not row_has_signal:
            continue

        # Contexto
        p = str(_coerce_scalar(row.get(col_pais))) if col_pais else pais_inferido
        d = str(_coerce_scalar(row.get(col_depto))) if col_depto else ""
        s = str(_coerce_scalar(row.get(col_sitio))) if col_sitio else ""

        # FILTRO por sitio: solo si la columna de sitio existe
        if col_sitio and not _is_nonempty_site(s):
            continue

        raw_mes = row.get(col_fecha_rep) if col_fecha_rep else (row.get(col_mesrep) if col_mesrep else None)
        m = month_label_from_value(raw_mes) or month_label_from_value(mes_inferido)

        p = p.strip() or pais_inferido
        d = d.strip()
        s = s.strip()
        m = m.strip() or month_label_from_value(mes_inferido)

        # Chequeo + posible error
        _add_metric(IND_TXML_CITA_VACIA, p, m, d, s, checks_add=1)
        v_exp = _coerce_scalar(row.get(col_cita_esperada))
        if pd.isna(v_exp) or str(v_exp).strip() == "":
            _add_metric(IND_TXML_CITA_VACIA, p, m, d, s, errors_add=1)
            errores_txml_cita.append({
                "País": p,
                "Departamento": d,
                "Sitio": s,
                "Mes de reporte": m,
                "Archivo": nombre_archivo,
                "ID expediente": str(_coerce_scalar(row.get(col_id))).strip() if col_id else "",
                "Modalidad de reporte": str(_coerce_scalar(row.get(col_modalidad))).strip() if col_modalidad else "",
                "Fecha de su última cita esperada": "",
                "Fila Excel": int(fila_base_txml + i),
                "Columna Excel": get_column_letter(col_cita_exp_idx + 1)
            })



def extraer_stage_txcurr_cohorte(
    xl: pd.ExcelFile, pais_inferido: str, mes_inferido: str, nombre_archivo: str, ruta_rel: str,
    stage_curr: List[Dict], stage_new: List[Dict], stage_rtt: List[Dict], stage_ml: List[Dict]
):
    """
    Extrae insumos para la conciliación trimestral TX_CURR.
    Reglas corregidas:
    1) TX_CURR solo usa la PRIMERA tabla (antes del bloque de periodicidad ARV).
    2) Para TX_NEW / TX_RTT / TX_ML el trimestre se infiere desde ruta/archivo
       porque varias plantillas traen metadatos superiores desactualizados.
    3) Para TX_NEW / TX_RTT / TX_ML se excluyen archivos con múltiples sitios
       dentro de la misma hoja, para evitar duplicación contra archivos sitio-específicos.
    """
    def _find_header_row(raw: pd.DataFrame, required_tokens: List[str], max_scan: int = 50) -> Optional[int]:
        lim = min(len(raw), max_scan)
        for r in range(lim):
            vals = [_norm(x) for x in raw.iloc[r].tolist()]
            if all(any(tok in v for v in vals) for tok in required_tokens):
                return r
        return None

    def _pick_col(cols, *patterns):
        return buscar_columna_multi(cols, *patterns)

    qfy_ctx, dt_ctx = _qfy_from_context_text(ruta_rel, nombre_archivo, mes_inferido)

    # ---- TX_CURR: solo primer bloque de desagregación por sexo/edad
    try:
        if "TX_CURR" in xl.sheet_names:
            raw = xl.parse("TX_CURR", header=None)
            hdr = _find_header_row(raw, ["sexo", "servicio"])
            if hdr is not None:
                stop_row = len(raw)
                for r in range(hdr + 1, len(raw)):
                    row_txt = " ".join(_norm(x) for x in raw.iloc[r].tolist() if _norm(x))
                    if "indicador:" in row_txt and "periodicidad de entrega" in row_txt:
                        stop_row = r
                        break

                df = raw.iloc[hdr + 1:stop_row].copy()
                df.columns = _dedupe_columns([str(c) for c in raw.iloc[hdr].tolist()])

                col_pais = "País" if "País" in df.columns else ("Pais" if "Pais" in df.columns else _pick_col(df.columns, "pais"))
                col_depto = _pick_col(df.columns, "departamento") or _pick_col(df.columns, "depto") or _pick_col(df.columns, "provincia")
                col_sitio = _pick_col(df.columns, "servicio", "salud") or _pick_col(df.columns, "sitio") or _pick_col(df.columns, "clinica")
                col_fecha = _pick_col(df.columns, "mes", "reporte") or _pick_col(df.columns, "fecha", "reporte")
                age_cols = _age_columns_tx_curr(list(df.columns))

                if col_sitio and age_cols:
                    tmp = df.copy()
                    for c in age_cols:
                        tmp[c] = pd.to_numeric(tmp[c], errors="coerce").fillna(0)

                    tmp["País"] = tmp[col_pais].apply(_canon_text) if col_pais else _canon_text(pais_inferido)
                    tmp["Departamento"] = tmp[col_depto].apply(_canon_text) if col_depto else ""
                    tmp["Sitio"] = tmp[col_sitio].apply(_canon_text)
                    tmp["FechaRepDT"] = tmp[col_fecha].apply(_parse_any_date) if col_fecha else dt_ctx
                    tmp["QFY"] = tmp["FechaRepDT"].apply(_qfy_from_date)

                    if qfy_ctx:
                        tmp.loc[tmp["QFY"].isna(), "QFY"] = qfy_ctx
                    if pd.notna(dt_ctx):
                        tmp.loc[tmp["FechaRepDT"].isna(), "FechaRepDT"] = dt_ctx

                    tmp["TX_CURR_row"] = tmp[age_cols].sum(axis=1)
                    tmp = tmp[(tmp["Sitio"] != "") & (tmp["QFY"].notna())]

                    if not tmp.empty:
                        grp = tmp.groupby(["País", "Sitio", "QFY", "FechaRepDT"], dropna=False, as_index=False).agg({
                            "TX_CURR_row": "sum",
                            "Departamento": lambda s: next((x for x in s.astype(str) if str(x).strip()), "")
                        })
                        grp["Archivo"] = nombre_archivo
                        stage_curr.extend(grp.to_dict("records"))
    except Exception:
        pass

    def _stage_event_sheet(sheet_name: str, target: List[Dict], include_rtt=False, include_ml=False):
        try:
            if sheet_name not in xl.sheet_names:
                return
            raw = xl.parse(sheet_name, header=None)
            hdr = _find_header_row(raw, ["id", "servicio"])
            if hdr is None:
                return

            df = raw.iloc[hdr + 1:].copy()
            df.columns = _dedupe_columns([str(c) for c in raw.iloc[hdr].tolist()])

            col_pais = "País" if "País" in df.columns else ("Pais" if "Pais" in df.columns else _pick_col(df.columns, "pais"))
            col_depto = _pick_col(df.columns, "departamento") or _pick_col(df.columns, "depto") or _pick_col(df.columns, "provincia")
            col_sitio = _pick_col(df.columns, "servicio", "salud") or _pick_col(df.columns, "sitio") or _pick_col(df.columns, "clinica")
            col_id = _pick_col(df.columns, "id")
            col_modalidad = _pick_col(df.columns, "modalidad", "reporte") or _pick_col(df.columns, "modalidad")

            if not (col_sitio and col_id):
                return

            tmp = pd.DataFrame({
                "País": df[col_pais].apply(_canon_text) if col_pais else _canon_text(pais_inferido),
                "Departamento": df[col_depto].apply(_canon_text) if col_depto else "",
                "Sitio": df[col_sitio].apply(_canon_text),
                "ID": df[col_id].astype(str).str.strip(),
            })
            tmp = tmp[(tmp["Sitio"] != "") & (tmp["ID"] != "") & (tmp["ID"].str.lower() != "nan")]

            if tmp.empty:
                return

            # Evita duplicar conteos cuando existe un archivo agregado con múltiples sitios
            if tmp["Sitio"].nunique(dropna=True) > 1:
                return

            tmp["QFY"] = qfy_ctx
            tmp["Archivo"] = nombre_archivo

            if include_rtt:
                if col_modalidad:
                    tmp["TipoRTT"] = df.loc[tmp.index, col_modalidad].astype(str).apply(
                        lambda x: "Traslado Recibido" if ("traslado" in _norm(x) and "recib" in _norm(x)) else "TX_RTT"
                    )
                else:
                    tmp["TipoRTT"] = "TX_RTT"

            if include_ml:
                tmp["Modalidad"] = df.loc[tmp.index, col_modalidad].astype(str).str.strip() if col_modalidad else ""

            tmp = tmp[tmp["QFY"].notna()].drop_duplicates()
            target.extend(tmp.to_dict("records"))
        except Exception:
            return

    _stage_event_sheet("TX_NEW", stage_new, include_rtt=False, include_ml=False)
    _stage_event_sheet("TX_RTT", stage_rtt, include_rtt=True, include_ml=False)
    _stage_event_sheet("TX_ML", stage_ml, include_rtt=False, include_ml=True)


def construir_validacion_txcurr_cohorte(
    stage_curr: List[Dict], stage_new: List[Dict], stage_rtt: List[Dict], stage_ml: List[Dict],
    errores_txcurr_cohorte: List[Dict], auditoria_txcurr_cohorte: List[Dict]
):
    if not stage_curr:
        return

    df_curr = pd.DataFrame(stage_curr)
    curr_by_date = df_curr.groupby(["País", "Sitio", "QFY", "FechaRepDT"], dropna=False, as_index=False).agg({
        "TX_CURR_row": "sum",
        "Departamento": lambda s: next((x for x in s.astype(str) if str(x).strip()), ""),
        "Archivo": lambda s: _join_unique_text(s)
    }).rename(columns={"TX_CURR_row": "TX_CURR_total", "Archivo": "Archivo(s) TX_CURR"})
    curr_final = (
        curr_by_date.sort_values(["País", "Sitio", "QFY", "FechaRepDT"])
        .groupby(["País", "Sitio", "QFY"], dropna=False, as_index=False)
        .tail(1)
        .reset_index(drop=True)
    )

    if stage_new:
        df_new = pd.DataFrame(stage_new).drop_duplicates(["País", "Sitio", "QFY", "ID"])
        new_group = df_new.groupby(["País", "Sitio", "QFY"], as_index=False).agg({
            "ID": "nunique",
            "Archivo": lambda s: _join_unique_text(s)
        }).rename(columns={"ID": "TX_NEW", "Archivo": "Archivo(s) TX_NEW"})
    else:
        new_group = pd.DataFrame(columns=["País", "Sitio", "QFY", "TX_NEW", "Archivo(s) TX_NEW"])

    if stage_rtt:
        df_rtt = pd.DataFrame(stage_rtt).drop_duplicates(["País", "Sitio", "QFY", "ID", "TipoRTT"])
        rtt_counts = df_rtt.groupby(["País", "Sitio", "QFY", "TipoRTT"], as_index=False)["ID"].nunique()
        rtt_files = df_rtt.groupby(["País", "Sitio", "QFY"], as_index=False)["Archivo"].agg(lambda s: _join_unique_text(s)).rename(columns={"Archivo": "Archivo(s) TX_RTT"})
        rtt_piv = rtt_counts.pivot_table(index=["País", "Sitio", "QFY"], columns="TipoRTT", values="ID", fill_value=0).reset_index()
        if "TX_RTT" not in rtt_piv.columns:
            rtt_piv["TX_RTT"] = 0
        if "Traslado Recibido" not in rtt_piv.columns:
            rtt_piv["Traslado Recibido"] = 0
        rtt_group = rtt_piv.merge(rtt_files, on=["País", "Sitio", "QFY"], how="left")
    else:
        rtt_group = pd.DataFrame(columns=["País", "Sitio", "QFY", "TX_RTT", "Traslado Recibido", "Archivo(s) TX_RTT"])

    if stage_ml:
        df_ml_total = pd.DataFrame(stage_ml).drop_duplicates(["País", "Sitio", "QFY", "ID"])
        ml_total = df_ml_total.groupby(["País", "Sitio", "QFY"], as_index=False).agg({
            "ID": "nunique",
            "Archivo": lambda s: _join_unique_text(s)
        }).rename(columns={"ID": "TX_ML total", "Archivo": "Archivo(s) TX_ML"})

        df_ml_mod = pd.DataFrame(stage_ml).drop_duplicates(["País", "Sitio", "QFY", "ID", "Modalidad"])
        ml_mod = df_ml_mod.groupby(["País", "Sitio", "QFY", "Modalidad"], as_index=False)["ID"].nunique()

        if not ml_mod.empty:
            ml_mod_piv = ml_mod.pivot_table(index=["País", "Sitio", "QFY"], columns="Modalidad", values="ID", fill_value=0).reset_index()
            rename_map = {c: f"TX_ML_{c}" for c in ml_mod_piv.columns if c not in ["País", "Sitio", "QFY"] and str(c).strip() != ""}
            ml_mod_piv = ml_mod_piv.rename(columns=rename_map)
            mod_desc = ml_mod.groupby(["País", "Sitio", "QFY"]).apply(
                lambda g: " | ".join([
                    f"{str(r['Modalidad']).strip()}: {int(r['ID'])}"
                    for _, r in g.iterrows()
                    if str(r['Modalidad']).strip()
                ])
            ).reset_index()
            mod_desc.columns = ["País", "Sitio", "QFY", "Modalidades TX_ML"]
            ml_group = ml_total.merge(mod_desc, on=["País", "Sitio", "QFY"], how="left").merge(ml_mod_piv, on=["País", "Sitio", "QFY"], how="left")
        else:
            ml_group = ml_total.copy()
            ml_group["Modalidades TX_ML"] = ""
    else:
        ml_group = pd.DataFrame(columns=["País", "Sitio", "QFY", "TX_ML total", "Archivo(s) TX_ML", "Modalidades TX_ML"])

    countries_with_quarters = curr_final.groupby("País")["QFY"].apply(lambda s: set(s.dropna())).to_dict()

    for _, rr in curr_final.iterrows():
        pais = rr["País"]
        sitio = rr["Sitio"]
        q_target = rr["QFY"]
        q_base = _prev_qfy(q_target)
        if not q_base or q_base not in countries_with_quarters.get(pais, set()):
            continue

        base_row = curr_final[(curr_final["País"] == pais) & (curr_final["Sitio"] == sitio) & (curr_final["QFY"] == q_base)]
        ng = new_group[(new_group["País"] == pais) & (new_group["Sitio"] == sitio) & (new_group["QFY"] == q_target)]
        rg = rtt_group[(rtt_group["País"] == pais) & (rtt_group["Sitio"] == sitio) & (rtt_group["QFY"] == q_target)]
        mg = ml_group[(ml_group["País"] == pais) & (ml_group["Sitio"] == sitio) & (ml_group["QFY"] == q_target)]

        base_val = int(base_row["TX_CURR_total"].iloc[0]) if not base_row.empty else 0
        base_fecha = base_row["FechaRepDT"].iloc[0] if not base_row.empty else pd.NaT
        base_arch = base_row["Archivo(s) TX_CURR"].iloc[0] if not base_row.empty else ""
        depto = rr.get("Departamento", "") if str(rr.get("Departamento", "")).strip() else (base_row["Departamento"].iloc[0] if (not base_row.empty and "Departamento" in base_row.columns) else "")

        tx_new = int(ng["TX_NEW"].iloc[0]) if not ng.empty else 0
        tx_rtt = int(rg["TX_RTT"].iloc[0]) if (not rg.empty and "TX_RTT" in rg.columns) else 0
        tras_rec = int(rg["Traslado Recibido"].iloc[0]) if (not rg.empty and "Traslado Recibido" in rg.columns) else 0
        tx_ml = int(mg["TX_ML total"].iloc[0]) if (not mg.empty and "TX_ML total" in mg.columns) else 0
        mods_ml = mg["Modalidades TX_ML"].iloc[0] if (not mg.empty and "Modalidades TX_ML" in mg.columns) else ""

        esperado = base_val + tx_new + tx_rtt + tras_rec - tx_ml
        real = int(rr["TX_CURR_total"])
        brecha = real - esperado

        audit_row = {
            "País": pais,
            "Departamento": depto,
            "Sitio": sitio,
            "Mes de reporte": q_target,
            "Trimestre base": q_base,
            "Trimestre comparado": q_target,
            "Fecha TX_CURR base": base_fecha.date() if pd.notna(base_fecha) else "",
            "Fecha TX_CURR comparado": rr["FechaRepDT"].date() if pd.notna(rr["FechaRepDT"]) else "",
            "Archivo(s) TX_CURR base": base_arch,
            "Archivo(s) TX_CURR comparado": rr.get("Archivo(s) TX_CURR", ""),
            "Archivo(s) TX_NEW": ng["Archivo(s) TX_NEW"].iloc[0] if (not ng.empty and "Archivo(s) TX_NEW" in ng.columns) else "",
            "Archivo(s) TX_RTT": rg["Archivo(s) TX_RTT"].iloc[0] if (not rg.empty and "Archivo(s) TX_RTT" in rg.columns) else "",
            "Archivo(s) TX_ML": mg["Archivo(s) TX_ML"].iloc[0] if (not mg.empty and "Archivo(s) TX_ML" in mg.columns) else "",
            "TX_CURR base": base_val,
            "TX_NEW": tx_new,
            "TX_RTT": tx_rtt,
            "Traslados recibidos": tras_rec,
            "TX_ML total": tx_ml,
            "Modalidades TX_ML": mods_ml,
            "TX_CURR esperado": esperado,
            "TX_CURR real": real,
            "Brecha (Real - Esperado)": brecha,
            "Estado": "Cuadra" if brecha == 0 else "No cuadra",
        }

        # agrega columnas de modalidades TX_ML desagregadas
        if not mg.empty:
            for c in mg.columns:
                if str(c).startswith("TX_ML_") and c not in audit_row:
                    audit_row[c] = int(mg[c].iloc[0]) if pd.notna(mg[c].iloc[0]) else 0

        if base_row.empty and real > 0:
            audit_row["Error identificado"] = "Sitio sin base en trimestre previo"
        elif brecha > 0:
            audit_row["Error identificado"] = "TX_CURR real mayor al esperado"
        elif brecha < 0:
            audit_row["Error identificado"] = "TX_CURR real menor al esperado"
        else:
            audit_row["Error identificado"] = "Cuadra"

        auditoria_txcurr_cohorte.append(audit_row)

        _add_metric(IND_TXCURR_COHORTE, pais, q_target, depto, sitio, checks_add=1)
        if brecha != 0:
            _add_metric(IND_TXCURR_COHORTE, pais, q_target, depto, sitio, errors_add=1)
            errores_txcurr_cohorte.append(audit_row.copy())

# ============================
# --------- PROCESO ----------

# ============================
if procesar:
    if not subida_multiple:
        st.warning("Primero carga archivos .xlsx o un .zip.")
        st.stop()

    entradas = []
    for up in subida_multiple:
        nombre = up.name; data = up.read()
        if nombre.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for info in zf.infolist():
                    if info.is_dir(): continue
                    if info.filename.lower().endswith(".xlsx") and not os.path.basename(info.filename).startswith("~$"):
                        entradas.append((os.path.basename(info.filename), zf.read(info.filename), info.filename))
        else:
            if nombre.lower().endswith(".xlsx") and not os.path.basename(nombre).startswith("~$"):
                entradas.append((os.path.basename(nombre), data, nombre))

    if not entradas:
        st.error("No se encontraron archivos .xlsx válidos.")
        st.stop()

    errores_numerador = []
    errores_txpvls = []
    errores_cd4 = []
    errores_tarv_gt = []
    errores_formato_fecha_diag = []
    errores_currq = []
    errores_iddup = []
    errores_sexo = []
    errores_txml_cita = []  # TX_ML
    errores_txcurr_cohorte = []  # Conciliación TX_CURR
    auditoria_txcurr_cohorte = []  # Auditoría tipo Auditoria_Sitio

    stage_curr = []
    stage_new = []
    stage_rtt = []
    stage_ml = []

    st.session_state.metrics_global = defaultdict(lambda: {"errors": 0, "checks": 0})
    st.session_state.metrics_by_pds = defaultdict(lambda: {"errors": 0, "checks": 0})

    progreso = st.progress(0.0, text="Procesando archivos…"); total = len(entradas)
    for idx, (nombre_archivo, data_bytes, ruta_rel) in enumerate(entradas, start=1):
        try:
            pais_inf, mes_inf = inferir_pais_mes(ruta_rel.replace("\\", "/"), default_pais, default_mes)
            xl = leer_excel_desde_bytes(nombre_archivo, data_bytes)
            procesar_tx_pvls_y_curr(xl, pais_inf, mes_inf, nombre_archivo, errores_numerador, errores_txpvls)
            procesar_hts_tst(xl, pais_inf, mes_inf, nombre_archivo,
                             errores_cd4, errores_formato_fecha_diag, errores_iddup,
                             errores_sexo, errores_tarv_gt)
            procesar_tx_ml_cita(xl, pais_inf, mes_inf, nombre_archivo, errores_txml_cita)  # TX_ML
            procesar_tx_curr_cuadros(xl, pais_inf, mes_inf, nombre_archivo, errores_currq)
            extraer_stage_txcurr_cohorte(xl, pais_inf, mes_inf, nombre_archivo, ruta_rel, stage_curr, stage_new, stage_rtt, stage_ml)
        except Exception as e:
            st.warning(f"⚠️ Error procesando {nombre_archivo}: {e}")
        progreso.progress(idx/total, text=f"Procesando {idx} de {total}…")

    st.session_state.df_num      = pd.DataFrame(errores_numerador)
    st.session_state.df_txpv     = pd.DataFrame(errores_txpvls)
    st.session_state.df_cd4      = pd.DataFrame(errores_cd4)
    st.session_state.df_tarv_gt  = pd.DataFrame(errores_tarv_gt)
    st.session_state.df_fdiag    = pd.DataFrame(errores_formato_fecha_diag)
    st.session_state.df_currq    = pd.DataFrame(errores_currq)
    st.session_state.df_iddup    = pd.DataFrame(errores_iddup)
    st.session_state.df_sexo     = pd.DataFrame(errores_sexo)
    construir_validacion_txcurr_cohorte(stage_curr, stage_new, stage_rtt, stage_ml, errores_txcurr_cohorte, auditoria_txcurr_cohorte)

    st.session_state.df_txml_cita = pd.DataFrame(errores_txml_cita)  # TX_ML
    st.session_state.df_txcurr_cohorte = pd.DataFrame(errores_txcurr_cohorte)  # Conciliación TX_CURR
    st.session_state.df_txcurr_auditoria = pd.DataFrame(auditoria_txcurr_cohorte)  # Auditoría TX_CURR

    # ===== Reordenar columnas: asegurar "Modalidad de reporte" justo tras "ID expediente" en TX_ML
    if not st.session_state.df_txml_cita.empty:
        cols = list(st.session_state.df_txml_cita.columns)
        after_col = "ID expediente" if "ID expediente" in cols else ("ID Expediente" if "ID Expediente" in cols else None)
        if after_col and "Modalidad de reporte" in cols:
            cols.remove("Modalidad de reporte")
            cols.insert(cols.index(after_col) + 1, "Modalidad de reporte")
            st.session_state.df_txml_cita = st.session_state.df_txml_cita[cols]

    # ===== NUEVO: Reordenar columnas en "ID (expediente) duplicado" para poner Sexo, Edad y Fecha del diagnóstico a la par de ID
    if not st.session_state.df_iddup.empty:
        cols = list(st.session_state.df_iddup.columns)
        after_col = "ID expediente" if "ID expediente" in cols else ("ID Expediente" if "ID Expediente" in cols else None)
        if after_col:
            # Quitar si ya están en otra posición
            for extra in ["Sexo", "Edad", "Fecha del diagnóstico"]:
                if extra in cols:
                    cols.remove(extra)
            insert_pos = cols.index(after_col) + 1
            if "Sexo" in st.session_state.df_iddup.columns:
                cols.insert(insert_pos, "Sexo"); insert_pos += 1
            if "Edad" in st.session_state.df_iddup.columns:
                cols.insert(insert_pos, "Edad"); insert_pos += 1
            if "Fecha del diagnóstico" in st.session_state.df_iddup.columns:
                cols.insert(insert_pos, "Fecha del diagnóstico")
            st.session_state.df_iddup = st.session_state.df_iddup[cols]

    st.session_state.processed = True
    st.success("Procesamiento completado. Ahora puedes filtrar al instante ✅")
("Procesamiento completado. Ahora puedes filtrar al instante ✅")

# ============================
# ------- INTERFAZ (LIVE) ----
# ============================
if not st.session_state.processed:
    st.info("Carga tus archivos y pulsa **Procesar**.")
    st.stop()

# Asegurar columnas base
for dfname in ["df_num","df_txpv","df_cd4","df_tarv_gt","df_fdiag","df_currq","df_iddup","df_sexo","df_txml_cita","df_txcurr_cohorte","df_txcurr_auditoria"]:
    df = st.session_state[dfname]
    if not isinstance(df, pd.DataFrame):
        st.session_state[dfname] = pd.DataFrame()
        continue
    if not df.empty:
        for col in ["País","Departamento","Sitio","Mes de reporte"]:
            if col not in df.columns:
                st.session_state[dfname][col] = ""

# Universo para segmentadores
df_all = pd.concat(
    [df for df in [
        st.session_state.df_num, st.session_state.df_txpv, st.session_state.df_cd4,
        st.session_state.df_tarv_gt, st.session_state.df_fdiag,
        st.session_state.df_currq, st.session_state.df_iddup, st.session_state.df_sexo,
        st.session_state.df_txml_cita, st.session_state.df_txcurr_cohorte  # TX_ML + Conciliación
    ] if isinstance(df, pd.DataFrame) and not df.empty],
    ignore_index=True
) if any([
    isinstance(st.session_state[k], pd.DataFrame) and not st.session_state[k].empty
    for k in ["df_num","df_txpv","df_cd4","df_tarv_gt","df_fdiag","df_currq","df_iddup","df_sexo","df_txml_cita","df_txcurr_cohorte"]
]) else pd.DataFrame(columns=["País","Departamento","Sitio","Mes de reporte"])

for c in ["País","Departamento","Sitio","Mes de reporte"]:
    if c in df_all.columns:
        df_all[c] = df_all[c].astype(str).str.replace("_", " ", regex=False).str.strip()

def _limpia_opts(vals):
    arr = []
    for v in vals:
        s = str(v).strip()
        if not s: 
            continue
        if s.lower() in ("desconocido", "nan"):
            continue
        arr.append(s)
    return ["Todos"] + sorted(set(arr))

def _on_change_pais():
    st.session_state.sel_depto = "Todos"
    st.session_state.sel_sitio = "Todos"

def _on_change_depto():
    st.session_state.sel_sitio = "Todos"

# 1) Segmentadores
seg = st.container(border=True)
with seg:
    st.subheader("🧊 Segmentadores")

    pais_opts = _limpia_opts(df_all["País"].dropna().tolist()) if "País" in df_all.columns else ["Todos"]
    if st.session_state.sel_pais not in pais_opts:
        st.session_state.sel_pais = "Todos"
    st.selectbox("País", pais_opts, key="sel_pais", on_change=_on_change_pais)

    df_p = df_all if st.session_state.sel_pais == "Todos" else df_all[df_all["País"] == st.session_state.sel_pais]
    depto_opts = _limpia_opts(df_p["Departamento"].dropna().tolist()) if "Departamento" in df_p.columns else ["Todos"]
    if st.session_state.sel_depto not in depto_opts:
        st.session_state.sel_depto = "Todos"
    st.selectbox("Departamento", depto_opts, key="sel_depto", on_change=_on_change_depto)

    df_pd = df_p if st.session_state.sel_depto == "Todos" else df_p[df_p["Departamento"] == st.session_state.sel_depto]
    sitio_opts = _limpia_opts(df_pd["Sitio"].dropna().tolist()) if "Sitio" in df_pd.columns else ["Todos"]
    if st.session_state.sel_sitio not in sitio_opts:
        st.session_state.sel_sitio = "Todos"
    st.selectbox("Sitio", sitio_opts, key="sel_sitio")

def _aplicar_filtro(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    m = pd.Series([True] * len(df))
    if st.session_state.sel_pais != "Todos":  m &= (df["País"].astype(str).str.replace("_", " ", regex=False).str.strip() == st.session_state.sel_pais)
    if st.session_state.sel_depto != "Todos": m &= (df["Departamento"].astype(str).str.replace("_", " ", regex=False).str.strip() == st.session_state.sel_depto)
    if st.session_state.sel_sitio != "Todos": m &= (df["Sitio"].astype(str).str.replace("_", " ", regex=False).str.strip() == st.session_state.sel_sitio)
    return df[m].copy()

df_num_f      = _aplicar_filtro(st.session_state.df_num)
df_txpv_f     = _aplicar_filtro(st.session_state.df_txpv)
df_cd4_f      = _aplicar_filtro(st.session_state.df_cd4)
df_tarv_gt_f  = _aplicar_filtro(st.session_state.df_tarv_gt)
df_fdiag_f    = _aplicar_filtro(st.session_state.df_fdiag)
df_currq_f    = _aplicar_filtro(st.session_state.df_currq)
df_iddup_f    = _aplicar_filtro(st.session_state.df_iddup)
df_sexo_f     = _aplicar_filtro(st.session_state.df_sexo)
df_txml_cita_f = _aplicar_filtro(st.session_state.df_txml_cita)  # TX_ML
df_txcurr_cohorte_f = _aplicar_filtro(st.session_state.df_txcurr_cohorte)  # Conciliación TX_CURR
df_txcurr_auditoria_f = _aplicar_filtro(st.session_state.df_txcurr_auditoria)  # Auditoría TX_CURR

# Métricas (adaptadas a la selección)
def _build_metrics_df_from_selection(sel_pais, sel_depto, sel_sitio):
    agg = defaultdict(lambda: {"errors": 0, "checks": 0})
    for (pais, depto, sitio, mes_rep, ind), v in st.session_state.metrics_by_pds.items():
        if (sel_pais == "Todos" or pais == sel_pais) and \
           (sel_depto == "Todos" or depto == sel_depto) and \
           (sel_sitio == "Todos" or sitio == sel_sitio):
            agg[ind]["errors"] += v["errors"]; agg[ind]["checks"] += v["checks"]

    df_global = (pd.DataFrame([
        {"Indicador": DISPLAY_NAMES.get(k, k), "Errores": v["errors"], "Chequeos": v["checks"], "% Error": _pct(v["errors"], v["checks"])}
        for k, v in agg.items()
    ]).sort_values("% Error", ascending=False)
    if agg else pd.DataFrame(columns=["Indicador","Errores","Chequeos","% Error"]))

    rows = []
    for (pais, depto, sitio, mes_rep, ind), v in st.session_state.metrics_by_pds.items():
        if (sel_pais == "Todos" or pais == sel_pais) and \
           (sel_depto == "Todos" or depto == sel_depto) and \
           (sel_sitio == "Todos" or sitio == sel_sitio):
            rows.append({
                "País": pais, "Departamento": depto, "Sitio": sitio, "Mes de reporte": mes_rep,
                "Indicador": DISPLAY_NAMES.get(ind, ind), "Errores": v["errors"], "Chequeos": v["checks"],
                "% Error": _pct(v["errors"], v["checks"])
            })
    df_group = pd.DataFrame(rows)
    if not df_group.empty:
        df_group = df_group[["País","Departamento","Sitio","Mes de reporte","Indicador","Errores","Chequeos","% Error"]]
        df_group = df_group.sort_values(["País","Departamento","Sitio","Indicador"])
    else:
        df_group = pd.DataFrame(columns=["País","Departamento","Sitio","Mes de reporte","Indicador","Errores","Chequeos","% Error"])
    return df_global, df_group

df_metricas_global_sel, df_metricas_por_mes_sel = _build_metrics_df_from_selection(
    st.session_state.sel_pais, st.session_state.sel_depto, st.session_state.sel_sitio
)

# 2) Resumen
res = st.container(border=True)
with res:
    st.subheader("⚫ *Resumen de errores por indicador*")
    c1, c2, c3, c4, c5 = st.columns(5)
    c6, c7, c8, c9, c10 = st.columns(5)
    c1.metric("*TX_PVLS (Num) > TX_PVLS (Den)*", len(df_num_f))
    c2.metric("*TX_PVLS (Den) > TX_CURR*", len(df_txpv_f))
    c3.metric("*CD4 vacío positivo*", len(df_cd4_f))
    c4.metric("*Fecha inicio TARV < Diagnóstico*", len(df_tarv_gt_f))
    c5.metric("*Fecha diag. mal formateada*", len(df_fdiag_f))
    c6.metric("*TX_CURR ≠ Dispensación_TARV*", len(df_currq_f))
    c7.metric("*ID duplicado - filas detectadas*", len(df_iddup_f))
    c8.metric("*Sexo inválido (HTS_TST)*", len(df_sexo_f))
    c9.metric("*TX_ML: Última cita esperada vacía*", len(df_txml_cita_f))
    c10.metric("*Conciliación TX_CURR trimestral*", len(df_txcurr_cohorte_f))

# 3) Indicadores – % de error (selección)
sel = st.container(border=True)
with sel:
    st.subheader("📊 *Porcentaje de errores por indicador*")
    cards = [IND_NUM_GT_DEN, IND_DEN_GT_CURR, IND_CD4_MISSING,
             IND_TARV_GT_DIAG, IND_DIAG_BAD_FMT, IND_CURR_Q1Q2_DIFF, IND_ID_DUPLICADO,
             IND_SEXO_INVALID, IND_TXML_CITA_VACIA, IND_TXCURR_COHORTE]  # TX_ML + Conciliación
    cols = st.columns(len(cards))
    sel_map = {row["Indicador"]: row for _, row in df_metricas_global_sel.iterrows()} if not df_metricas_global_sel.empty else {}
    for col, key in zip(cols, cards):
        name = DISPLAY_NAMES[key]
        v = sel_map.get(name, {"Errores":0, "Chequeos":0, "% Error":0})
        col.metric(label=name, value=f"{v.get('% Error',0)}%", delta=f"{v.get('Errores',0)} / {v.get('Chequeos',0)} (error/chequeos)")

# 4) Detalle por indicador
det = st.container(border=True)
with det:
    # --- Estilo: tabs con scroll horizontal (Detalle por indicador)
    st.markdown("""
        <style>
        /* Hace desplazable horizontalmente la barra de pestañas */
        .stTabs [data-baseweb="tab-list"]{
        overflow-x: auto !important;
        overflow-y: hidden;
        white-space: nowrap;
        gap: .5rem;
        padding-bottom: 2px;
        scrollbar-width: thin;  /* Firefox */
    }
        .stTabs [data-baseweb="tab"]{
        flex: 0 0 auto;         /* evita que las tabs se encojan y permite el scroll */
    }
        </style>
            """, unsafe_allow_html=True)

    st.subheader("🔎 *Detalle por indicador*")

    tab_specs = [
        ("TX_PVLS (Num) > TX_PVLS (Den)", df_num_f,   "— Sin diferencias de Numerador > Denominador —"),
        ("TX_PVLS (Den) > TX_CURR",       df_txpv_f,  "— Sin casos Denominador > TX_CURR —"),
        ("CD4 vacío positivo",            df_cd4_f,   "— Sin positivos con CD4 vacío —"),
        ("Fecha inicio TARV < Fecha diagnóstico", df_tarv_gt_f, "— Sin casos Fecha inicio TARV anterior al diagnóstico —"),
        ("Formato fecha diagnóstico",     df_fdiag_f, "— Sin problemas de formato de fecha —"),
        ("TX_CURR ≠ Dispensación_TARV",   df_currq_f, "— TX_CURR = Dispensación_TARV en la selección —"),
        ("ID (expediente) duplicado",     df_iddup_f, "— Sin IDs (expediente) duplicados —"),
        ("Sexo inválido (HTS_TST)",       df_sexo_f,  "— Sin filas con sexo inválido —"),
        ("TX_ML: Última cita esperada vacía", df_txml_cita_f, "— Sin filas con 'Última cita esperada' vacía —"),  # TX_ML
        ("Conciliación TX_CURR trimestral", df_txcurr_cohorte_f, "— Sin diferencias en la conciliación trimestral TX_CURR —"),
        ("Auditoría Sitio TX_CURR", df_txcurr_auditoria_f, "— Sin filas de auditoría TX_CURR —"),
    ]

    tabs = st.tabs([title for title, _, _ in tab_specs])
    for t, (_, df_, empty_note) in zip(tabs, tab_specs):
        with t:
            show_df_or_note(df_, empty_note, height=340)

# 5) Métricas de calidad (adaptadas al filtro)
met = st.container(border=True)
with met:
    st.subheader("📈 *Resumen de porcentajes de error por indicador y desglose por país*")
    gc1, gc2 = st.columns([1.2, 2])
    with gc1:
        st.markdown("**Métricas – Selección actual**")
        show_df_or_note(df_metricas_global_sel, "— Sin métricas para la selección —", height=260)
    with gc2:
        st.markdown("**Desglose por Mes – Selección**")
        show_df_or_note(df_metricas_por_mes_sel, "— Sin desglose para la selección —", height=260)

# ============================
# ---------- DESCARGA --------
# ============================
def exportar_excel_resultados(errores_dict, df_metricas_global: pd.DataFrame, df_metricas_group: pd.DataFrame) -> bytes:
    # Soporta uno o varios campos por hoja a resaltar (lista o string)
    config_resaltado = {
        "Numerador > Denominador": "Numerador",
        "Denominador > TX_CURR": "Denominador (PVLS)",
        "CD4 vacío positivo": ["CD4 Basal", "Motivo de no CD4"],
        "Fecha TARV < Diagnóstico (HTS)": ["Fecha diagnóstico", "Fecha inicio TARV"],
        "Formato fecha diagnóstico": "Fecha del diagnóstico de la prueba",
        "TX_CURR ≠ Dispensación_TARV": "Diferencia (TX_CURR - Disp_TARV)",
        "ID (expediente) duplicado": "ID expediente",
        "Sexo inválido (HTS_TST)": "Sexo (valor encontrado)",
        "TX_ML: Última cita esperada vacía": "Fecha de su última cita esperada",  # TX_ML
        "Conciliación TX_CURR trimestral": ["Brecha (Real - Esperado)", "TX_CURR esperado", "TX_CURR real"],
        "Auditoria_Sitio TX_CURR": ["Brecha (Real - Esperado)", "TX_CURR esperado", "TX_CURR real"],
    }
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used = set()
        resumen = pd.DataFrame({
            "Tipo de Error": list(errores_dict.keys()),
            "Total": [len(df) for df in errores_dict.values()]
        })
        resumen.to_excel(writer, index=False, sheet_name=_safe_sheet_name("Resumen", used))
        for nombre, df in errores_dict.items():
            if df is not None and not df.empty:
                df.to_excel(writer, index=False, sheet_name=_safe_sheet_name(nombre, used))
        if df_metricas_global is not None and not df_metricas_global.empty:
            df_metricas_global.to_excel(writer, index=False, sheet_name=_safe_sheet_name("Metricas Globales Seleccion", used))
        if df_metricas_group is not None and not df_metricas_group.empty:
            df_metricas_group.to_excel(writer, index=False, sheet_name=_safe_sheet_name("Metricas por Mes", used))

    buffer.seek(0); wb = load_workbook(buffer)
    rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for nombre, df in errores_dict.items():
        if df is None or df.empty: continue
        target = None
        for ws in wb.worksheets:
            if ws.title.lower().startswith(_safe_sheet_name(nombre, set()).lower()[:5]):
                target = ws.title; break
        if not target: continue
        ws = wb[target]
        campo_rojo = config_resaltado.get(nombre)
        if campo_rojo:
            campos = campo_rojo if isinstance(campo_rojo, list) else [campo_rojo]
            for camp in campos:
                if camp in df.columns:
                    col_idx = list(df.columns).index(camp) + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).fill = rojo
        ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

full_dict = {
    "Numerador > Denominador": st.session_state.df_num,
    "Denominador > TX_CURR": st.session_state.df_txpv,
    "CD4 vacío positivo": st.session_state.df_cd4,
    "Fecha TARV < Diagnóstico (HTS)": st.session_state.df_tarv_gt,
    "Formato fecha diagnóstico": st.session_state.df_fdiag,
    "TX_CURR ≠ Dispensación_TARV": st.session_state.df_currq,
    "ID (expediente) duplicado": st.session_state.df_iddup,
    "Sexo inválido (HTS_TST)": st.session_state.df_sexo,
    "TX_ML: Última cita esperada vacía": st.session_state.df_txml_cita,  # TX_ML
    "Conciliación TX_CURR trimestral": st.session_state.df_txcurr_cohorte,
    "Auditoria_Sitio TX_CURR": st.session_state.df_txcurr_auditoria,
}

rows_metrics_global = [
    {"Indicador": DISPLAY_NAMES[k], "Errores": v["errors"], "Chequeos": v["checks"], "% Error": _pct(v["errors"], v["checks"]) }
    for k, v in st.session_state.metrics_global.items()
]
df_metricas_global_all = (
    pd.DataFrame(rows_metrics_global).sort_values("% Error", ascending=False)
    if rows_metrics_global else pd.DataFrame(columns=["Indicador","Errores","Chequeos","% Error"])
)

rows_all = []
for (pais, depto, sitio, mes_rep, ind), v in st.session_state.metrics_by_pds.items():
    rows_all.append({
        "País": pais, "Departamento": depto, "Sitio": sitio, "Mes de reporte": mes_rep,
        "Indicador": DISPLAY_NAMES[ind], "Errores": v["errors"], "Chequeos": v["checks"],
        "% Error": _pct(v["errors"], v["checks"])
    })

df_metricas_por_mes_all = pd.DataFrame(rows_all)
if not df_metricas_por_mes_all.empty:
    df_metricas_por_mes_all = df_metricas_por_mes_all[["País","Departamento","Sitio","Mes de reporte","Indicador","Errores","Chequeos","% Error"]]
    df_metricas_por_mes_all = df_metricas_por_mes_all.sort_values(["País","Departamento","Sitio","Indicador"])

bytes_excel_full = exportar_excel_resultados(full_dict, df_metricas_global_all, df_metricas_por_mes_all)

filt_dict = {
    "Numerador > Denominador": df_num_f,
    "Denominador > TX_CURR": df_txpv_f,
    "CD4 vacío positivo": df_cd4_f,
    "Fecha TARV < Diagnóstico (HTS)": df_tarv_gt_f,
    "Formato fecha diagnóstico": df_fdiag_f,
    "TX_CURR ≠ Dispensación_TARV": df_currq_f,
    "ID (expediente) duplicado": df_iddup_f,
    "Sexo inválido (HTS_TST)": df_sexo_f,
    "TX_ML: Última cita esperada vacía": df_txml_cita_f,  # TX_ML
    "Conciliación TX_CURR trimestral": df_txcurr_cohorte_f,
    "Auditoria_Sitio TX_CURR": df_txcurr_auditoria_f,
}
bytes_excel_filt = exportar_excel_resultados(filt_dict, df_metricas_global_sel, df_metricas_por_mes_sel)

fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
# Fix menor para evitar NameError si 'pais' no existe en este contexto:
pais = st.session_state.get("sel_pais", default_pais)  # (no afecta validaciones)

dl = st.container(border=True)
with dl:
    cdl1, cdl2 = st.columns(2)
    with cdl1:
        st.download_button(
            "⬇️ Descargar Excel (COMPLETO)",
            data=bytes_excel_full,
            file_name=f"Errores_validaciones_{pais}_{fecha_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with cdl2:
        st.download_button(
            "⬇️ Descargar Excel (FILTRADO)",
            data=bytes_excel_filt,
            file_name=f"Errores_validaciones_ {pais}_{fecha_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )




